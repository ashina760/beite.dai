# workflow.py
import os
import shutil
import pandas as pd
from openpyxl import  Workbook
from excel_handler.processor import ExcelProcessor
from excel_handler.utils import check_dates_in_dict, check_past_dates
from settings import (TITLE_COLUMNS, EXPECTED_TITLES,MANDATORY_CELLS, MANDATORY_COLUMN,DATE_COLUMN, ID_COLUMN,FILL_VALUES,MIN_COL, MAX_COL)
from settings import REFERENCE_PATH,KEY_COLUMNS_IN_A, KEY_COLUMNS_IN_B, VALUE_COLUMN_IN_B, TARGET_COLUMN_IN_A,DOWNLOADS_PATH,CELLS_LIST
from openpyxl.utils import column_index_from_string
from excel_handler.utils import get_latest_file,format_date


def validate_excel_data(processor: ExcelProcessor) -> dict:
    """
    执行一系列校验，如标题、空单元格、历史日期等。
    如果发现错误，返回包含错误信息及详情的字典。
    """
    errors = {}

    if processor.has_multiple_sheets():
        errors["sheets"] = {"sheet not only one"}

    title_result = processor.is_title_valid(TITLE_COLUMNS, expected_values=EXPECTED_TITLES)
    if title_result:
        errors["title"] = {title_result}

    processor.delete_empty_rows(MANDATORY_COLUMN)

    if processor.is_cell_empty(MANDATORY_CELLS):
        errors["cell_check"] = {"cell_check is empty"}

    empty_cells = processor.find_empty_cells(MIN_COL, MAX_COL)
    if empty_cells:
        errors["empty_cells"] = empty_cells

    
    id_date_tuple = processor.get_column_dates_with_colD(DATE_COLUMN, ID_COLUMN)

    b = ExcelProcessor(REFERENCE_PATH)

    delivery_date_dict = b.get_column_based_dict()

    unmatched = check_dates_in_dict(id_date_tuple, delivery_date_dict)
    if unmatched:
        errors["notfound"] = {"date": unmatched}

    past_dates = check_past_dates(id_date_tuple)
    if past_dates:
        errors["past_date"] = {"date": past_dates}

    b.close()
    
    return errors


def create_upload_data_from_processor(processor: ExcelProcessor, save_dir):
    """
    从 ExcelProcessor 类对象生成上传用的 Excel 文件。

    参数:
        processor (ExcelProcessor): 拥有 sheet, min_row, sheet_name 属性的类实例
        save_dir (str): 保存目录
        fill_values (dict): 要填充的列和默认值

    返回:
        str: 保存的文件路径
    """

    sheet = processor.sheet
    min_row = processor.min_row
    sheet_name = processor.sheet_name

    headers = [
        "T", "仕入先コード", "センターコード", "指定納期", "担当者コード", "決裁区分", "決裁番号", "発注残管理",
        "商品コード", "発注数量", "明細備考1", "明細備考2", "決裁営業", "お客様", "伝票備考"
    ]

    os.makedirs(save_dir, exist_ok=True)

    wb_new = Workbook()
    sheet_new = wb_new.active
    sheet_new.title = sheet_name
    sheet_new.append(headers)

    for row in sheet.iter_rows(min_row=min_row, values_only=True):
        new_row = [
            None, row[2], row[3], row[7], None, None, None, None,
            row[4], row[6], None, None, None, None, row[10]
        ]
        sheet_new.append(new_row)

    # 日期格式化（第4列）
    for row in sheet_new.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            format_date(cell)

    # 转字符串列
    for col in [2, 3, 9, 10, 15]:
        for row in sheet_new.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                if cell.value is not None:
                    cell.value = str(cell.value)

    # 补填默认值
    for row_idx in range(2, sheet_new.max_row + 1):
        for col, val in FILL_VALUES.items():
            sheet_new.cell(row=row_idx, column=col, value=val)

    # 删除发注数量为空或为 0 的行（第10列）
    rows_to_delete = [
        row[0].row for row in sheet_new.iter_rows(min_row=2, min_col=10, max_col=10)
        if row[0].value in [None, 0, "0"]
    ]
    for row_idx in reversed(rows_to_delete):
        sheet_new.delete_rows(row_idx)

    save_path = os.path.join(save_dir, "nagashikomi.xlsx")
    wb_new.save(save_path)
    return save_path

# def transfer_matching_data(a: ExcelProcessor, b: ExcelProcessor):
#     """
#     对比两个表的数据，如果 key 匹配，则把 B 表指定列值写入 A 表指定列。

#     参数:
#         processor_a: ExcelProcessor，主表处理器（写入）
#         processor_b: ExcelProcessor，参考表处理器（查找）
#         key_columns_a: list[str]，A表构造key的列名
#         key_columns_b: list[str]，B表构造key的列名（顺序需与A一致）
#         value_column_b: str，B表中需提取的列（如 'F'）
#         write_column_a: str，A表中要写入的列（如 'M'）

#     返回:
#         int: 成功匹配写入的行数
#     """
    
#     key_columns_a=["C", "D", "E", "G", "H","K"],
#     key_columns_b=["J", "B", "Z", "AE", "E", "X"],
#     value_column_b="E",
#     write_column_a="M",

#     # 提取 B 表数据，生成 key -> value 映射
#     key_to_value = {}
#     for row in range(b.min_row, b.max_row + 1):
#         key = tuple(b.sheet[f"{col}{row}"].value for col in key_columns_b)
#         value = b.sheet[f"{value_column_b}{row}"].value
#         key_to_value[key] = value

#     write_count = 0

#     # 在 A 表中查找 key 并写入
#     for row in range(a.min_row, a.max_row + 1):
#         key = tuple(a.sheet[f"{col}{row}"].value for col in key_columns_a)
#         if key in key_to_value:
#             a.sheet[f"{write_column_a}{row}"].value = key_to_value[key]
#             write_count += 1

#     return write_count


def match_and_fill_from_csv(processor: ExcelProcessor):
    """
    在 A 表中，根据指定列组合 key，在 B (CSV) 表中查找匹配项，如果找到则将指定列的值写入 A 表目标列。
    
    参数:
        processor: ExcelProcessor 实例 (处理 A 表)
        csv_path: CSV 文件路径 (B 表)
        key_columns_in_a: List[str]，A 表中参与 key 的列名，如 ['C', 'D', 'E']
        key_columns_in_b: List[str]，CSV 中对应的列名，如 ['col1', 'col2', 'col3']
        value_column_in_b: str，CSV 中要写入 A 表的值所在的列名，如 'F'
        target_column_in_a: str，写入 A 表的目标列名，如 'M'
    """
    csv_path = get_latest_file(DOWNLOADS_PATH)

    df_b = pd.read_csv(csv_path, encoding="cp932", dtype=str).fillna("")  # 读取并填空字符串，避免 NaN 干扰
    df_b["key"] = df_b.apply(lambda row: build_clean_key(row, KEY_COLUMNS_IN_B), axis=1)
    key_value_dict = dict(zip(df_b["key"], df_b[VALUE_COLUMN_IN_B]))  # 假设你要记录的是 F 列的值
    processor.get_min_max_row()
    max_row = processor.max_row
    processor.convert_column_to_yyyymmdd("H")
    target_col_idx = column_index_from_string(TARGET_COLUMN_IN_A)
    # print(key_value_dict)
    # 遍历 A 表的每一行，构造 key 并匹配写入值
    for row in range(2, max_row + 1):  # 从第2行开始跳过表头
        key_parts = [str(processor.sheet[f"{col}{row}"].value).strip() for col in KEY_COLUMNS_IN_A]
        full_key = ''.join(key_parts)
        full_key = full_key.replace(" ", "").replace("\u3000", "").replace("\n", "")  # ✨ 清理空格
        # print(f"构造的 key: {full_key}")

        if full_key in key_value_dict:
            processor.sheet.cell(row=row, column=target_col_idx, value=key_value_dict[full_key])
    return csv_path

def build_clean_key(row, key_columns):
    def normalize(val):
        if pd.isna(val):
            return ""
        val = str(val).strip().replace('\u3000', '').replace('\n', '')
        try:
            if 'e' in val.lower():
                val = format(float(val), '.0f')  # 去除科学计数法
        except:
            pass
        return val
    return ''.join([normalize(row[col]) for col in key_columns])

def move_csv_to_folder(csv_path, new_folder_path):
    # 确保目标文件夹存在，不存在则创建
    if not os.path.exists(new_folder_path):
        os.makedirs(new_folder_path)
    
    # 获取CSV文件的文件名
    file_name = os.path.basename(csv_path)
    
    # 构建目标路径
    new_path = os.path.join(new_folder_path, file_name)
    
    # 移动文件
    shutil.move(csv_path, new_path)

    return new_path

def tantou_name(processor: ExcelProcessor):
    name = processor.get_cell_values_from_workbook(CELLS_LIST)
    return name

def save_excel_object(processor: ExcelProcessor, save_path: str = None):
    """
    保存 Excel 对象的工作簿。

    参数:
        excel_obj: 包含 workbook 和 file_path 属性的对象。
        save_path (str, optional): 保存路径。如果未提供，则使用默认路径。
    """
    if save_path is None:
        filename = os.path.basename(processor.file_path)
        directory = os.path.dirname(processor.file_path)
        save_path = os.path.join(directory, f"NEW_{filename}")

    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    processor.save(save_path)
    return save_path
    